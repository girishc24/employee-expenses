from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenRefreshView
from .serializers import CustomTokenRefreshSerializer
import os
from datetime import date
import random
import string
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.core.mail import send_mail
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.response  import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.views import APIView
from rest_framework.exceptions import AuthenticationFailed
from . serializers import *
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.core.cache import cache
from django.contrib.auth.hashers import make_password
from .models import *
from django.utils.dateparse import parse_date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
from rest_framework import generics
from django.db.models import Sum
import logging
import json
from PIL import Image
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
import sys
from django.core.files.base import ContentFile

from django.db.models import Q
import zipfile
from django.template.loader import render_to_string
from weasyprint import HTML
from xhtml2pdf import pisa
import requests
from django.core.mail import EmailMessage
from .services import handle_new_subscription, update_subscription_status


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        try:
            data = super().validate(attrs)
            refresh = self.get_token(self.user)

            data['refresh'] = str(refresh)
            data['access'] = str(refresh.access_token)
            data['access_expiry'] = refresh.access_token.payload['exp']
            return data
        except AuthenticationFailed:
            raise AuthenticationFailed({'error': 'Invalid credentials. Please try again.'})


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer



class CustomTokenRefreshView(TokenRefreshView):
    serializer_class = CustomTokenRefreshSerializer

def home(request):
    return render(request, "index.html")
# def home(request):
#     return HttpResponse("Welcome to BizmiTT")

def privacy(request):
    return render(request, "privacy-policy.html")

def termsandcondition(request):
    return render(request, "terms-condition.html")

def refundpolicy(request):
    return render(request, "refund-policy.html")

@api_view(['POST'])
def adduser(request):
    if request.method == 'POST':
        user_serializer = UserCreateSerializer(data=request.data)
        employee_serializers = EmployeeSerializers(data=request.data)
        

        if user_serializer.is_valid(raise_exception=True) and employee_serializers.is_valid(raise_exception=True):
            user = user_serializer.save()
            employee = employee_serializers.save(user=user)
            response_data = {
                'user': user_serializer.data,
                'employee': employee_serializers.data
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        return Response({'error': user_serializer.errors, 'error': employee_serializers.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    else:
        return Response({'error': 'GET Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)



@receiver(post_save, sender=User)
def create_default_categories(sender, instance, created, **kwargs):
    if created:
        def add_days(d, days):
            return d + timedelta(days=days)
        today = date.today()
        three_months_later = add_days(today, 15)
        plan = Subscriptions.objects.get(id=1)
        #print(f"Plan ID {plan}")
        deleted_user = DeletedAccount.objects.filter(
            emailid=instance.username
        ).exists()

        if not deleted_user:
            
            subscription = Usersubscription.objects.create(
                user=instance,
                sub_plan=plan,
                start_date=today,
                end_date=three_months_later,
                amt=plan.price,
                status='Current Plan',
                available=50
            )
            print(f"Created subscription: {subscription}")
        else:
            print(f"User {instance.username} is in DeletedAccount; no subscription created.")

        
        
        travel_photo = 'employeeapp/images/travel.png'
        accomodation_photo = 'employeeapp/images/Accomodations.png'
        food_photo = 'employeeapp/images/Food.png'
        items_photo = 'employeeapp/images/MaterialPurchase.png'
        miscellaneous_photo = 'employeeapp/images/miscellaneous.png'

        travel_category = Category.objects.create(name='Travel', user=instance, photo=travel_photo)
        accomodation_category = Category.objects.create(name='Accommodation', user=instance, photo=accomodation_photo)
        food_category = Category.objects.create(name='Food Expenses', user=instance, photo=food_photo)
        items_category = Category.objects.create(name='Material Purchased', user=instance, photo=items_photo)
        miscellaneous_category = Category.objects.create(name='Miscellaneous', user=instance, photo=miscellaneous_photo)


        Subcategory.objects.create(category=travel_category, name='Flight', user=instance)
        Subcategory.objects.create(category=travel_category, name='Taxi', user=instance)
        Subcategory.objects.create(category=travel_category, name='Bus', user=instance)
        Subcategory.objects.create(category=travel_category, name='Train', user=instance)
        Subcategory.objects.create(category=travel_category, name='Rapido', user=instance)
        Subcategory.objects.create(category=travel_category, name='Ola', user=instance)
        Subcategory.objects.create(category=travel_category, name='Uber', user=instance)
        Subcategory.objects.create(category=travel_category, name='Namma Yatri', user=instance)
        Subcategory.objects.create(category=travel_category, name='Bike', user=instance)
        Subcategory.objects.create(category=travel_category, name='Others', user=instance)


        Subcategory.objects.create(category=accomodation_category, name='Accommodation', user=instance)

        
        Subcategory.objects.create(category=items_category, name='Material Purchased', user=instance)

        
        Subcategory.objects.create(category=food_category, name='Breakfast', user=instance)
        Subcategory.objects.create(category=food_category, name='Lunch', user=instance)
        Subcategory.objects.create(category=food_category, name='Snacks', user=instance)
        Subcategory.objects.create(category=food_category, name='Dinner', user=instance)

        
        Subcategory.objects.create(category=miscellaneous_category, name='Medical Expenses', user=instance)
        Subcategory.objects.create(category=miscellaneous_category, name='Cell phone', user=instance)
        Subcategory.objects.create(category=miscellaneous_category, name='Health & Wellness', user=instance)
        Subcategory.objects.create(category=miscellaneous_category, name='Learning & Development', user=instance)
        Subcategory.objects.create(category=miscellaneous_category, name='Entertainment', user=instance)
        Subcategory.objects.create(category=miscellaneous_category, name='Refreshment', user=instance)

@api_view(['POST'])
def check_email_phone(request):
    if request.method == 'POST':
        email = request.data.get('email', None)
        phone = request.data.get('phone', None)
        
        if email and phone:
            if User.objects.filter(email=email).exists():
                return Response({'error': 'Email already exists'}, status=status.HTTP_400_BAD_REQUEST)
            elif Employee.objects.filter(phone=phone).exists():
                return Response({'error': 'Phone Number already exists'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                otp = ''.join(random.choices(string.digits, k=6))
                subject = 'Welcome to BizmITT! Verify Your Email'
                message = f'''
                Welcome to BizmITT - your go-to app for
                managing expenses effortlessly!
                To get started, please verify your email
                using the OTP below:
                Your OTP: {otp}
                Looking forward to helping you streamline
                your reimbursements!
                Best regards,
                InnoThrive Technologies
                '''

                send_mail(
                    subject,
                    message,
                    'Email Verification',  
                    [email],
                    fail_silently=False,
                )
                return Response({'message': 'OTP sent to your email', 'OTP': otp}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Both email and phone must be provided'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'error': 'Method Not Allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['POST'])
def validate_otp(request):
    if request.method == 'POST':
        otp_entered = request.data.get('otp_entered')
        otp_generated = request.data.get('otp_generated')  
        
        print(f'otp_entered: {otp_entered}, otp_generated: {otp_generated}')
        
        if otp_entered == otp_generated:
            return Response({'message': 'OTP validated successfully'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid OTP'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'error': 'Method Not Allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)



@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def viewprofile(request):
    if request.method == 'GET':
        user_id = request.user.id
        try:
            user = User.objects.get(id=user_id)
            user_serializer = UserSerializer(user)

            try:
                employee = Employee.objects.get(user=user)
                emp_serializer = EmployeeSerializers(employee, context={'request': request})
                employeedata = {
                    'user': user_serializer.data,
                    'employee': emp_serializer.data
                }
                return Response(employeedata, status=status.HTTP_200_OK)
            except Employee.DoesNotExist:
                return Response({"error": "Employee not found for this user."}, status=status.HTTP_404_NOT_FOUND)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
    else:
        return Response({'error': 'Invalid Method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def editprofile(request):
    if request.method == 'PUT':
        user_instance = request.user
        user_data = request.data
        user_serializer = UserEditSerializersnew(instance=user_instance, data=user_data)
        employee_data = request.data
        employee_instance = user_instance.employee  
        employee_serializer = EmployeeEditSerializers(instance=employee_instance, data=employee_data, context={'request': request})

        if user_serializer.is_valid(raise_exception=True) and employee_serializer.is_valid(raise_exception=True):
            user_serializer.save()
            employee_serializer.save()

            response_data = {
                'user': user_serializer.data,
                'employee': employee_serializer.data
            }
            return Response(response_data, status=status.HTTP_200_OK)

        return Response({'error': user_serializer.errors, 'error': employee_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    else:
        return Response({'error': 'GET Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    
        

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def categories(request):
    if request.method == 'GET':
        user_id = request.user.id
        categories = Category.objects.filter(user_id=user_id).order_by('-id')
        category_serializer = CategorySerializer(categories, many=True, context={'request': request})

        response_data = {
            "category": category_serializer.data,
        }

        return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def subcategories(request, pk):
    if request.method == 'GET':
        subcategories = Subcategory.objects.filter(category_id=pk)
        subcategory_serializer = SubcategorySerializer(subcategories, many=True)
        return Response(subcategory_serializer.data, status=status.HTTP_200_OK)

# def compress_image(image):
#     try:
#         im = Image.open(image)
#         output = BytesIO()

#         # Convert image mode to 'RGB' to handle all cases
#         if im.mode in ('RGBA', 'LA'):
#             im = im.convert('RGB')

#         # Resize the image
#         max_size = (1024, 1024)
#         im.thumbnail(max_size, Image.LANCZOS)

#         # Compress and save the image
#         im.save(output, format='JPEG', quality=20, optimize=True)

#         # Reset the buffer position to the beginning
#         output.seek(0)

#         # Debug logs
#         original_size = len(image.read())
#         compressed_size = len(output.getvalue())
#         print(f"Original size: {original_size} bytes")
#         print(f"Compressed size: {compressed_size} bytes")

#         return InMemoryUploadedFile(
#             output,
#             'ImageField',
#             f"{image.name.split('.')[0]}.jpg",
#             'image/jpeg',
#             compressed_size,
#             None
#         )
#     except Exception as e:
#         print(f"Error compressing image: {e}")
#         raise

def compress_image_to_size(image_file, target_size_kb, output_format='JPEG'):
    target_size_bytes = target_size_kb * 1024  # Convert KB to bytes
    quality = 85  # Start with an initial quality level
    buffer = BytesIO()
    
    try:
        # Open the image file
        image = Image.open(image_file)

        # Convert the image to RGB mode if necessary
        if image.mode in ("RGBA", "P"):
            image = image.convert("RGB")

        while True:
            buffer.seek(0)
            buffer.truncate()
            
            # Save the image to the buffer with the current quality level and output format
            image.save(buffer, format=output_format, quality=quality, optimize=True)

            # Check the size of the image
            size = buffer.tell()

            # If the image is within the target size, break the loop
            if size <= target_size_bytes or quality <= 5:
                break
            
            # Reduce the quality to further compress the image
            quality -= 5

        # Create a ContentFile from the buffer
        compressed_file = ContentFile(buffer.getvalue())

        # Define a new file name for the compressed image with the appropriate extension
        compressed_file_name = f"{os.path.splitext(image_file.name)[0]}_compressed.{output_format.lower()}"
        
        return compressed_file, compressed_file_name
    except Exception as e:
        raise Exception(f"Image compression failed: {str(e)}")

logger = logging.getLogger(__name__)
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Expenses(APIView):
    def post(self, request):
        user = request.user
        
        
        current_plan = Usersubscription.objects.filter(
            user_id=user.id,
            status="Current Plan",
            end_date__gte=timezone.now().date()
        ).order_by('-end_date').first()

        
        if not current_plan:
            current_plan = Usersubscription.objects.filter(
                user_id=user.id,
                status="Queued Plan"
            ).order_by('-start_date').first()
            
            if current_plan:
                
                current_plan.status = "Current Plan"
                current_plan.save()
            else:
                
                return Response({'error': 'No active or valid subscription found.'}, status=status.HTTP_400_BAD_REQUEST)

        
        current_date = timezone.now().date()

        if current_plan.end_date < current_date:
            update_subscription_status(current_plan)
            return Response({'error': 'Your subscription has expired.'}, status=status.HTTP_400_BAD_REQUEST)

        if current_plan.available <= 0:
            update_subscription_status(current_plan)
            return Response({'error': 'Your usage limit has been reached. Please upgrade to a new plan.'}, status=status.HTTP_400_BAD_REQUEST)

        files = request.FILES.getlist('document')
        serializer = ExpenseSerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            try:
                expense = serializer.save()
                for file in files:
                    if file.content_type.startswith('image'):
                        # Compress the image to the desired size (e.g., 200 KB) and format (JPEG)
                        compressed_file, compressed_file_name = compress_image_to_size(file, 200, output_format='JPEG')  # Adjust target size in KB and format
                        if compressed_file:
                            expense_doc = ExpenseDocument.objects.create(
                                expense=expense,
                                document=compressed_file
                            )
                            expense_doc.document.save(compressed_file_name, compressed_file)
                        else:
                            return Response({'error': 'File compression failed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    else:
                        ExpenseDocument.objects.create(expense=expense, document=file)

                
                current_plan.available -= 1
                current_plan.credits_used += 1
                current_plan.save()

                
                update_subscription_status(current_plan)

                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request, pk=None):  
        if pk is not None:  
            expense = Expense.objects.filter(id=pk).first()
            if expense:
                expense_serializer = ExpenseSerializerview(expense, context={'request': request})
                return Response(expense_serializer.data, status=status.HTTP_200_OK)
            else:
                return Response({'error': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)
        else:  
            user = request.user
            expenses = Expense.objects.filter(user=user, archived=False).order_by('-id')
            
            start_date = request.query_params.get('startdate')
            end_date = request.query_params.get('enddate')
            categories = request.query_params.get('categories')

            
            if start_date:
                start_date = parse_date(start_date)
                if start_date:
                    expenses = expenses.filter(expense_date__gte=start_date)

            
            if end_date:
                end_date = parse_date(end_date)
                if end_date:
                    expenses = expenses.filter(expense_date__lte=end_date)

            
            if categories:
                categories_list = [category.strip() for category in categories.split(',')]
                expenses = expenses.filter(category__name__in=categories_list)
            
            
            expenses_serializer = ExpenseSerializerNew(expenses, many=True, context={'request': request})
            return Response(expenses_serializer.data, status=status.HTTP_200_OK)
            

         
    
    
    def put(self, request, pk):
        try:
            expense = Expense.objects.get(id=pk, user=request.user)
        except Expense.DoesNotExist:
            return Response({"error": "Expense not found"}, status=status.HTTP_404_NOT_FOUND)

        
        serializer = ExpenseSerializerEdit(expense, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        
        document_actions = json.loads(request.data.get('document_actions', '[]'))
        
        for action in document_actions:
            if action['action'] == 'delete':
                ExpenseDocument.objects.filter(id=action['id'], expense=expense).delete()
            elif action['action'] == 'keep':
                pass  

        
        new_files = request.FILES.getlist('document')
        for file in new_files:
            ExpenseDocument.objects.create(expense=expense, document=file)

        
        updated_expense = ExpenseSerializer(expense).data
        updated_documents = ExpenseDocumentSerializer(expense.documents.all(), many=True).data

        return Response({
            "expense": updated_expense,
            "documents": updated_documents
        }, status=status.HTTP_200_OK)
    
    
    def delete(self, request, pk):
        try:
            logger.debug(f"Delete request received for expense ID: {pk}")
            expense = get_object_or_404(Expense, id=pk)
            expense.delete()
            logger.info(f"Expense ID {pk} deleted successfully.")
            return Response({'message': 'Expense deleted successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            logger.error(f"Error occurred while deleting expense ID {pk}: {str(e)}")
            return Response({'error': 'An error occurred while deleting the expense'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])   
class Resetdata(generics.GenericAPIView):
    def get_queryset(self):
        user = self.request.user
        return Expense.objects.filter(user=user)

    def delete(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'error': 'No data found'}, status=status.HTTP_404_NOT_FOUND)
        
        
        deleted_count, _ = queryset.delete()
        existing_category = [
        'Refreshment', 'Entertainment', 'Learning & Development', 'Health & Wellness', 
        'Cell phone', 'Medical Expenses', 'Dinner', 'Snacks', 'Lunch', 'Breakfast', 
        'Material Purchased', 'Accommodation', 'Others', 'Namma Yatri', 'Uber', 'Ola', 
        'Rapido', 'Bike', 'Train', 'Bus', 'Taxi', 'Flight'
        ]
        
        new_category = []
        user = request.user  
        sub_category = Subcategory.objects.filter(user=user)  

        for sub in sub_category:
            if sub.name not in existing_category:  
                new_category.append(sub)

        
        for new in new_category:
            new.delete()

        return Response({'message': f'{deleted_count} Expense Data Reset Successfully'}, status=status.HTTP_204_NO_CONTENT)

@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Archive(APIView):   
    def get(self, request, pk=None):  
        if pk is not None:  
            expense = Expense.objects.filter(id=pk).first()
            if expense:
                expense_serializer = ExpenseSerializerview(expense, context={'request': request})
                return Response(expense_serializer.data, status=status.HTTP_200_OK)
            else:
                return Response({'message': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)
        else:  
            user = request.user
           
            expenses = Expense.objects.filter(user=user, archived=True)
            expenses_serializer = ExpenseSerializerNew(expenses, many=True, context={'request': request})

            
            start_date = request.query_params.get('startdate')
            end_date = request.query_params.get('enddate')
            categories = request.query_params.get('categories')

            
            if start_date:
                start_date = parse_date(start_date)
                if start_date:
                    expenses = expenses.filter(expense_date__gte=start_date)

            
            if end_date:
                end_date = parse_date(end_date)
                if end_date:
                    expenses = expenses.filter(expense_date__lte=end_date)

            
            if categories:
                categories_list = [category.strip() for category in categories.split(',')]
                expenses = expenses.filter(category__name__in=categories_list)

            
            total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0

            dashboard_data = {
                'archived_total': total_expense,
                'expense': ExpenseSerializerNew(expenses, many=True, context={'request': request}).data
            }

            return Response(dashboard_data, status=status.HTTP_200_OK)
    
    def delete(self, request, pk):
        try:
            expense = get_object_or_404(Expense, id=pk)
            expense.delete()
            return Response({'message': 'Expense deleted successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({'error': 'An error occurred while deleting the expense'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Viewreport(APIView):
    def get(self, request):
        user = request.user
        employee = Employee.objects.get(user=user)  
        expenses = Expense.objects.filter(user=user, archived=False).order_by('-id')
        documents = ExpenseDocument.objects.filter(expense__in=expenses)
        for document in documents:
            print(document)

        
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')
        categories = request.query_params.get('categories')
        export_format = request.query_params.get('export', 'json')

       
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                expenses = expenses.filter(expense_date__gte=start_date)
        
        
        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                expenses = expenses.filter(expense_date__lte=end_date)
        
        
        if categories:
            categories_list = [category.strip() for category in categories.split(',')]
            expenses = expenses.filter(category__name__in=categories_list)

        if start_date is None:
            start_date = expenses.values_list('expense_date', flat=True).order_by('expense_date').first()

        if end_date is None:
            end_date = expenses.values_list('expense_date', flat=True).order_by('-expense_date').first()


        
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0

        domain = request.build_absolute_uri('/').rstrip('/')

        if export_format == 'excel':
            excel_buffer = self.export_to_excel(expenses, domain, employee, start_date, end_date, total_expense)
            pdf_buffer = self.export_to_pdf(expenses, employee, start_date, end_date, total_expense, request)
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                zip_file.writestr('expense_report.xlsx', excel_buffer.getvalue())
                zip_file.writestr('expense_report.pdf', pdf_buffer.getvalue())
            
            zip_buffer.seek(0)

            
            self.send_report_zip_via_email(employee.user.email, zip_buffer, employee, start_date, end_date)

            
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=expense_reports.zip'

            return response
        else:
            
            expenses_serializer = ExpenseSerializerNew(expenses, many=True, context={'request': request})
            return Response(expenses_serializer.data, status=status.HTTP_200_OK)

    def export_to_excel(self, expenses, domain, employee, start_date, end_date, total_expense):
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Report"
        
        
        alignment = Alignment(horizontal="left", vertical="top")

        
        ws['A1'] = f'Employee Name :'
        ws['B1'] = f'{employee.user.get_full_name()}'
        ws['A2'] = f'Employee ID : '
        ws['B2'] = f'{employee.emp_id}'
        ws['A3'] = f'Designation : '
        ws['B3'] = f'{employee.designation}'
        ws['A4'] = f'Department :'
        ws['B4'] = f'{employee.department}'
        ws['E1'] = 'Expense From Date:'
        ws['F1'] = str(start_date) if start_date else 'N/A'
        ws['E2'] = 'Expense End Date:'
        ws['F2'] = str(end_date) if end_date else 'N/A'
        ws['E3'] = 'Expenses Submit Date:'
        ws['F3'] = str(timezone.now().date())
        ws['E4'] = 'Expenses Total Amount:'
        ws['F4'] = str(total_expense)

        for row in range(1, 5):
            for col in ['A', 'E']:
                ws[f'{col}{row}'].alignment = alignment

       
        headers = ['SL.no', 'Created Date', 'Expense Date', 'Main Category', 'Sub Category', 
                   'Payment Mode', 'Proof Type', 'Remarks', 'Amount', 'Proof Image']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)

        
        for row, expense in enumerate(expenses, start=7):
            ws.cell(row=row, column=1, value=row - 6)  
            ws.cell(row=row, column=2, value=expense.created_date)
            ws.cell(row=row, column=3, value=expense.expense_date)
            ws.cell(row=row, column=4, value=expense.category.name)
            ws.cell(row=row, column=5, value=expense.subcategory.name if expense.subcategory else '')
            ws.cell(row=row, column=6, value=expense.payment)
            ws.cell(row=row, column=7, value=expense.proof)
            ws.cell(row=row, column=8, value=expense.note)
            ws.cell(row=row, column=9, value=expense.amount)
            
            document_urls = [self.get_file_url(doc.document, domain) for doc in expense.documents.all()]
            ws.cell(row=row, column=10, value="\n".join(document_urls))

        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def export_to_pdf(self, expenses, employee, start_date, end_date, total_expense,request):
        document = ExpenseDocument.objects.filter(expense__in=expenses)

        print(expenses)  
        print(document)  

        context = {
            'employee': employee,
            'expenses': expenses,
            'start_date': start_date,
            'end_date': end_date,
            'total_expense': total_expense,
            'document':document
        }

        html_string = render_to_string('reportpdf.html', context)
        pdf_buffer = BytesIO()
        #pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)


        return pdf_buffer


    def send_report_zip_via_email(self, email, zip_buffer, employee, start_date, end_date):
        subject = f"Your BizmITT Expense Report is Ready ({start_date} to {end_date})"
        message = f"""
        Hi {employee.user.get_full_name()},
        
        Your requested expense report has been successfully generated (From {start_date} to {end_date}). 
        Please find the attached report for your records.
        
        Important Note: If you would like to use this as a final report, please don't forget to move these expenses to the archive in our app. 
        Archiving will help keep your expense tracking and analytics accurate for active expenses.
        
        Best regards,
        InnoThrive Technologies
        """

        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email],
        )

        # Attach the ZIP file
        email.attach('expense_reports.zip', zip_buffer.getvalue(), 'application/zip')

        # Send the email
        email.send()

    def get_file_url(self, file, domain=None):
        if file and hasattr(file, 'name'):
            return f"{domain}{settings.MEDIA_URL}{file.name}" if domain else f"{settings.MEDIA_URL}{file.name}"
        return ''


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Archivereport(APIView):
    def get(self, request):
        user = request.user
        employee = Employee.objects.get(user=user)
        expenses = Expense.objects.filter(user=user).order_by('-id')
        documents = ExpenseDocument.objects.filter(expense__in=expenses)
        for document in documents:
            print(document)
        

       
        # Get query parameters
        start_date = request.query_params.get('startdate')
        end_date = request.query_params.get('enddate')
        categories = request.query_params.get('categories')
        export_format = request.query_params.get('export', 'json')

        # Filter by start date
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                expenses = expenses.filter(expense_date__gte=start_date)

        # Filter by end date
        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                expenses = expenses.filter(expense_date__lte=end_date)

        # Filter by categories
        if categories:
            categories_list = [category.strip() for category in categories.split(',')]
            expenses = expenses.filter(category__name__in=categories_list)
        
        if start_date is None:
            start_date = expenses.values_list('expense_date', flat=True).order_by('expense_date').first()

        if end_date is None:
            end_date = expenses.values_list('expense_date', flat=True).order_by('-expense_date').first()


        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0

        # Update filtered expenses to archived=True
        with transaction.atomic():
            updated_count = expenses.update(archived=True, archived_date=timezone.now())
        
        domain = request.build_absolute_uri('/').rstrip('/')

        if export_format == 'excel':
            excel_buffer = self.export_to_excel(expenses, domain, employee, start_date, end_date, total_expense)
            pdf_buffer = self.export_to_pdf(expenses, employee, start_date, end_date, total_expense, request)

            # Create ZIP file containing both Excel and PDF
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                zip_file.writestr('archived_expense_report.xlsx', excel_buffer.getvalue())
                zip_file.writestr('archived_expense_report.pdf', pdf_buffer.getvalue())

            zip_buffer.seek(0)

            # Send the email with the ZIP file attached
            self.send_email_with_attachment(user, start_date, end_date, zip_buffer)

            # Respond with the ZIP file as an attachment if needed
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=archived_expense_reports.zip'

            return response
        else:
            # Serialize the filtered and updated data
            expenses_serializer = ExpenseSerializerNew(expenses, many=True, context={'request': request})
            return Response({
                'message': f'{updated_count} expenses have been archived.',
                'archived_expenses': expenses_serializer.data
            }, status=status.HTTP_200_OK)

    def export_to_excel(self, expenses, domain, employee, start_date, end_date, total_expense):
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        # Set alignment for the top section
        alignment = Alignment(horizontal="left", vertical="top")

        # Add custom headings at the top of the sheet with employee details
        ws['A1'] = f'Employee Name :'
        ws['B1'] = f'{employee.user.get_full_name()}'
        ws['A2'] = f'Employee ID : '
        ws['B2'] = f'{employee.emp_id}'
        ws['A3'] = f'Designation : '
        ws['B3'] = f'{employee.designation}'
        ws['A4'] = f'Department :'
        ws['B4'] = f'{employee.department}'
        ws['E1'] = 'Expense From Date:'
        ws['F1'] = str(start_date) if start_date else 'N/A'
        ws['E2'] = 'Expense End Date:'
        ws['F2'] = str(end_date) if end_date else 'N/A'
        ws['E3'] = 'Expenses Submit Date:'
        ws['F3'] = str(timezone.now().date())
        ws['E4'] = 'Expenses Total Amount:'
        ws['F4'] = str(total_expense)

        for row in range(1, 5):
            for col in ['A', 'E']:
                ws[f'{col}{row}'].alignment = alignment

        # Write table headers
        headers = ['SL.no', 'Created Date', 'Expense Date', 'Main Category', 'Sub Category',
                   'Payment Mode', 'Proof Type', 'Remarks', 'Amount', 'Proof Image']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)

        # Write data
        for row, expense in enumerate(expenses, start=7):
            ws.cell(row=row, column=1, value=row - 6)  # Serial number
            ws.cell(row=row, column=2, value=expense.created_date)
            ws.cell(row=row, column=3, value=expense.expense_date)
            ws.cell(row=row, column=4, value=expense.category.name)
            ws.cell(row=row, column=5, value=expense.subcategory.name if expense.subcategory else '')
            ws.cell(row=row, column=6, value=expense.payment)
            ws.cell(row=row, column=7, value=expense.proof)
            ws.cell(row=row, column=8, value=expense.note)
            # Fetch and concatenate document URLs
            document_urls = [self.get_file_url(doc.document, domain) for doc in expense.documents.all()]
            ws.cell(row=row, column=10, value="\n".join(document_urls))

        # Create a BytesIO buffer to save the workbook to
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def export_to_pdf(self, expenses, employee, start_date, end_date, total_expense, request):
        document = ExpenseDocument.objects.filter(expense__in=expenses)

        print("expenses Documents")  
        print(document)  

        context = {
            'employee': employee,
            'expenses': expenses,
            'start_date': start_date,
            'end_date': end_date,
            'total_expense': total_expense,
            'document':document
        }

        html_string = render_to_string('reportpdf.html', context)
        pdf_buffer = BytesIO()
        #pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)


        return pdf_buffer

    def send_email_with_attachment(self, user, start_date, end_date, zip_buffer):
        subject = f'Your Expenses Have Been Archived - BizmITT Report {start_date} to {end_date}'
        message = f'''
            Hi {user.first_name},

            Your expenses have been successfully archived
            from {start_date} to {end_date}. Please find
            the attached report for your reference.

            Important Note: The expenses you have archived
            cannot be unarchived and will not appear in your
            active expenses list. You can view these expenses
            in the "Archived Expenses" section of the app.
            After 30 days, the proof attachment images will
            be automatically deleted from our database and
            will not be available for future reference in our
            application. Kindly keep the attachments with this mail
            for future records.

            Best regards,
            InnoThrive Technologies
        '''
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [user.email]

        # Create the email with attachment
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=from_email,
            to=recipient_list
        )

        # Attach the ZIP file
        email.attach('archived_expense_reports.zip', zip_buffer.getvalue(), 'application/zip')

        # Send the email
        email.send()

    def get_file_url(self, file, domain=None):
        if file and hasattr(file, 'name'):
            return f"{domain}{settings.MEDIA_URL}{file.name}" if domain else f"{settings.MEDIA_URL}{file.name}"
        return ''


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Addsubcategory(APIView):
    def post(self, request, pk):
        user = request.user.id
        try:
            category = Category.objects.get(id=pk)
        except Category.DoesNotExist:
            return Response({"error": "Category not found."}, status=status.HTTP_404_NOT_FOUND)
        
        # Make a mutable copy of request.data
        data = request.data.copy()
        data['category'] = category.id
        data['user'] = user
        
        serializer = AddsubcategorySerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                

@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Subscriptiondetails(APIView):
    def get(self, request):
        sub = Subscriptions.objects.exclude(id=1).order_by('id')
        subscriptions_serializer = SubscriptionSerialixer(sub, many=True)
        return Response(subscriptions_serializer.data, status=status.HTTP_200_OK)
    
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class UserSubscriptionView(APIView):
    def get(self, request):
        user = request.user.id
        users=Usersubscription.objects.filter(user=user)
        subscriptio_serializer=UsersubscriptionSerializer(users, many=True)
        return Response(subscriptio_serializer.data, status=status.HTTP_200_OK)


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
# class Subscriptionrenewal(APIView):
#     def post(self, request):
#         user = request.user
#         sub_plan_id = request.data.get('sub_plan_id')

#         # Fetch the subscription plan
#         try:
#             sub_plan = Subscriptions.objects.get(id=sub_plan_id)
#         except Subscriptions.DoesNotExist:
#             return Response({"error": "Subscription plan not found"}, status=404)

#         # Calculate the duration in days
#         duration_days = sub_plan.duration 

#         # Check if the user has any active or previous subscriptions
#         existing_subscription = Usersubscription.objects.filter(user=user).exists()

#         if not existing_subscription:
#             status = "Current Plan"
#             start_date = timezone.now().date()
#         else:
#             active_subscription = Usersubscription.objects.filter(
#                 user=user,
#                 end_date__gte=timezone.now().date()
#             ).first()

#             if active_subscription:
#                 # Preserve the status of the current plan
#                 if active_subscription.status in ["Current Plan", "Queued Plan"]:
#                     start_date = active_subscription.end_date + timedelta(days=1)
#                     status = "Queued Plan"
#                 else:
#                     start_date = timezone.now().date()
#                     status = "Current Plan"
#             else:
#                 start_date = timezone.now().date()
#                 status = "Current Plan"

#         # Calculate the new end date for the new subscription
#         new_end_date = start_date + timedelta(days=duration_days)

#         razorpay_order_id = request.data.get('razorpay_order_id')
#         razorpay_payment_id = request.data.get('razorpay_payment_id')
#         amt = sub_plan.price

#         # Create a new subscription record
#         Usersubscription.objects.create(
#             user=user,
#             sub_plan=sub_plan,
#             start_date=start_date,
#             end_date=new_end_date,
#             razorpay_order_id=razorpay_order_id,
#             razorpay_payment_id=razorpay_payment_id,
#             amt=amt,
#             status=status,
#             available=sub_plan.available,
#         )

#         return Response({"message": "Subscription renewed successfully", "new_end_date": new_end_date})
class Subscriptionrenewal(APIView):
    def post(self, request):
        user = request.user
        sub_plan_id = request.data.get('sub_plan_id')

        # Fetch the subscription plan
        try:
            sub_plan = Subscriptions.objects.get(id=sub_plan_id)
        except Subscriptions.DoesNotExist:
            return Response({"error": "Subscription plan not found"}, status=status.HTTP_404_NOT_FOUND)
        
       
        # Calculate the duration in days
        duration_days = sub_plan.duration 

        # Check for any active or previous subscriptions
        active_subscription = Usersubscription.objects.filter(
            user=user,
            status="Current Plan",
            end_date__gte=timezone.now().date()
        ).first()

        if active_subscription:
            if active_subscription.status == "Credits Reached":
                # If the current plan has reached its credits, activate the queued plan if it exists
                queued_plan = Usersubscription.objects.filter(
                    user=user,
                    status="Queued Plan"
                ).first()

                if queued_plan:
                    queued_plan.status = "Current Plan"
                    queued_plan.save()
                
                # Queue the new plan since the queued plan has been activated
                start_date = queued_plan.end_date + timedelta(days=1)
                subscription_status = "Queued Plan"
            else:
                # If the current plan is still valid but not expired, queue the new plan
                start_date = active_subscription.end_date + timedelta(days=1)
                subscription_status = "Queued Plan"
        else:
            # If there's no active subscription, start the new one today
            start_date = timezone.now().date()
            subscription_status = "Current Plan"

        # Calculate the new end date for the new subscription
        new_end_date = start_date + timedelta(days=duration_days)

        razorpay_order_id = request.data.get('razorpay_order_id')
        razorpay_payment_id = request.data.get('razorpay_payment_id')
        amt = sub_plan.price

        # Create a new subscription record
        Usersubscription.objects.create(
            user=user,
            sub_plan=sub_plan,
            start_date=start_date,
            end_date=new_end_date,
            razorpay_order_id=razorpay_order_id,
            razorpay_payment_id=razorpay_payment_id,
            amt=amt,
            status=subscription_status,  # Status determined based on logic
            available=sub_plan.available,
        )

        return Response({"message": "Subscription renewed successfully", "new_end_date": new_end_date})


@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Subscriptioncondition(APIView):
    def post(self, request):
        user = request.user
        sub_plan_id = request.data.get('sub_plan_id')

        try:
            sub_plan = Subscriptions.objects.get(id=sub_plan_id)
        except Subscriptions.DoesNotExist:
            return Response({"error": "Subscription plan not found"}, status=status.HTTP_404_NOT_FOUND)
        
        queued_plan_exists = Usersubscription.objects.filter(
            user=user,
            status="Current Plan"
        ).exists()

        if queued_plan_exists:
            return Response({"error": "You already have a Current plan. Please use the current plan before purchasing a new one."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"message": "You can make a New renewal", "sub_plan_id": sub_plan_id}, status=status.HTTP_200_OK)
    
@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def dashboardanalysis(request):
    if request.method == 'GET':
        user_id = request.user.id
        
        # Calculate total expenses
        total_expense = Expense.objects.filter(user_id=user_id, archived=False).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Initialize category data with all categories set to 0
        categories = Category.objects.all()
        category_data = {category.name: 0 for category in categories}
        
        # Calculate expenses per category
        category_expenses = Expense.objects.filter(user_id=user_id, archived=False).values('category__name').annotate(total_amount=Sum('amount'))
        
        # Update category data with actual expenses
        for category in category_expenses:
            category_name = category['category__name']
            category_total = category['total_amount']
            category_data[category_name] = category_total
        
        # Prepare response data
        expenses = {
            "amount": total_expense
        }
        dashboard_data = {
            'expenses': expenses,
            'category': category_data
        }
        
        return Response(dashboard_data, status=status.HTTP_200_OK)
    else:
        return Response({'error': 'Invalid Method'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    



@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class VerifyEmail(APIView):
    def get(self, request):
        email = request.user.email
        if email:
            otp = ''.join(random.choices(string.digits, k=6))
            #cache.set(f'otp_{request.user.id}', otp, timeout=300)  # store OTP in cache for 5 minutes
            print(otp)
            send_mail(
                'BizmITT Account Deletion Request - OTP Verification',
                f'''
                Hi {request.user.first_name},

                We have received a request to delete your
                BizmITT account. To proceed with the 
                deletion, please verify your request
                by entering the OTP below:

                Your OTP: {otp}

                This OTP is valid for the next 10 minutes. If
                you did not request to delete your account, 
                please ignore this email and your account
                will remain active.

                Please note that once your account is 
                deleted, all your data will be permanently
                removed and cannot be recovered.

                Thank you for using BizmITT.

                Best regards,
                InnoThrive Technologies
                ''',
                'no-reply@innothrive.com',
                [email],
                fail_silently=False,
            )

            return Response({'otp': otp, 'message': 'OTP sent to your email'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Email not provided'}, status=status.HTTP_400_BAD_REQUEST)

@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class VerifyOTP(APIView):
    def post(self, request):
        otp_entered = request.data.get('otp_entered')
        otp_generated = request.data.get('otp_generated')

        print(f'otp_entered: {otp_entered}, otp_generated: {otp_generated}')
        user = request.user
        phone_number = user.employee.phone

        if otp_entered and otp_generated and otp_entered == otp_generated:
            DeletedAccount.objects.create(emailid=user.email, phoneno=phone_number)
            request.user.delete()
            return Response({'message': 'Account deleted successfully'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid OTP'}, status=status.HTTP_400_BAD_REQUEST)

@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
class Helpview(APIView):
    def get(self, request):
        help=Help.objects.all()
        help_serializer= HelpSerializer(help, many=True)
        return Response(help_serializer.data, status=status.HTTP_200_OK)


class Privacypolicy(APIView):
    def get(self, request):
        privacy_policies = PrivacyPolicy.objects.all()
        privacy_serializer= PrivacyPolicySerializer(privacy_policies, many=True)
        return Response(privacy_serializer.data, status=status.HTTP_200_OK)

class Termsandcondition(APIView):
    def get(self, request):
        privacy_policies = Termscondition.objects.all()
        privacy_serializer= TermsSerializer(privacy_policies, many=True)
        return Response(privacy_serializer.data, status=status.HTTP_200_OK)

@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])   
class Faqview(APIView):
    def get(self, request):
        faq = Faq.objects.all()
        faq_serializer = FaqSerializer(faq, many=True)
        return Response(faq_serializer.data, status=status.HTTP_200_OK)
    
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])   
class Razorpaykeyview(APIView):
    def get(self, request):
        razorpay = Razorpaykey.objects.all()
        razorpay_serializer = RazorpaykeySerializer(razorpay, many=True)
        return Response(razorpay_serializer.data, status=status.HTTP_200_OK)



    

@api_view(['POST'])
def forgotpassword(request):
    email = request.data.get('email')  
    if email:
        if User.objects.filter(email=email).exists():
            otp = ''.join(random.choices(string.digits, k=6))
            subject = 'Reset Your BizmITT Password'
            message = f'''
            We received a request to reset your BizmITT password. Use the OTP below to proceed:

            Your OTP: {otp}

            If you didn't request a password reset, please ignore this email.

            Best regards,
            InnoThrive Technologies
            '''

            # Send the email
            send_mail(
                subject,
                message,
                'Forgot Password OTP',  
                [email],
                fail_silently=False,
            )
            return Response({'message': 'OTP sent to your email', 'OTP': otp, 'email': email}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'No Account Found in this Mail ID'}, status=status.HTTP_404_NOT_FOUND)
    else:
        return Response({'error': 'Email not provided'}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['POST'])
def forgotpasswordotpvalidate(request):
    if request.method == 'POST':
        otp_entered = request.data.get('otp_entered')
        otp_generated = request.data.get('otp_generated')
        password = request.data.get('password')
        email = request.data.get('email')  
        user = User.objects.get(email=email)
        if otp_entered == otp_generated:
            user.password = make_password(password)
            user.save()
            return Response({'message': 'OTP Validated Successfully & Password Updated Successfully'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid OTP'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response({'error': 'Method Not Allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    
def pdfreport(request):
    # Filter expenses for the user with user_id 83
    expenses = Expense.objects.filter(user_id=85)
    document = ExpenseDocument.objects.filter(expense__in=expenses)
    print("expenses")
    print(expenses)
    print("document")
    print(document)
    # Pass the filtered expenses to the template
    context = {
        'expenses': expenses,
        'document': document
    }
    
    return render(request, "reportpdf.html", context)